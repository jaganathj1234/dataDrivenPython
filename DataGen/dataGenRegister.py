import openpyxl


def dataGeneratorRegister():
    workbook = openpyxl.load_workbook("./ConfigurationFile/reg_doc.xlsx")
    sheet = workbook['Sheet1']
    Row = sheet.max_row
    li_reg = []
    li1 = []
    for i in range(1, Row + 1):
        li1 = []
        firstname = sheet.cell(i, 1)
        lastname = sheet.cell(i, 2)
        email = sheet.cell(i, 3)
        password = sheet.cell(i, 4)
        confirmpass = sheet.cell(i, 5)
        li1.insert(0, firstname.value)
        li1.insert(1, lastname.value)
        li1.insert(2, email.value)
        li1.insert(3, password.value)
        li1.insert(4, confirmpass.value)
        li_reg.insert(i - 1, li1)
    return li_reg
