import openpyxl


def dataGeneratorLogin():
    workbook = openpyxl.load_workbook("./ConfigurationFile/reg_doc.xlsx")
    sheet = workbook['Sheet2']
    Row = sheet.max_row
    li_log = []
    li2 = []
    for i in range(1, Row + 1):
        li2 = []
        username = sheet.cell(i, 1)
        password = sheet.cell(i, 2)
        li2.insert(0, username.value)
        li2.insert(1, password.value)
        li_log.insert(i - 1, li2)
    return li_log
