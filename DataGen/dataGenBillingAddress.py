import openpyxl


def dataGeneratorBillingAddress():
    workbook = openpyxl.load_workbook("./ConfigurationFile/reg_doc.xlsx")
    sheet = workbook['Sheet3']
    Row = sheet.max_row
    li_ba = []
    li3 = []
    for i in range(1, Row + 1):
        li3 = []
        username = sheet.cell(i, 1)
        password = sheet.cell(i, 2)
        company = sheet.cell(i, 3)
        city = sheet.cell(i, 4)
        address1 = sheet.cell(i, 5)
        address2 = sheet.cell(i, 6)
        zip = sheet.cell(i, 7)
        phone = sheet.cell(i, 8)
        fax = sheet.cell(i, 9)
        li3.insert(0, username.value)
        li3.insert(1, password.value)
        li3.insert(2, company.value)
        li3.insert(3, city.value)
        li3.insert(4, address1.value)
        li3.insert(5, address2.value)
        li3.insert(6, zip.value)
        li3.insert(7, phone.value)
        li3.insert(8, fax.value)
        li_ba.insert(i - 1, li3)
    return li_ba
